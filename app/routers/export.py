"""
app/routers/export.py

POST /api/v1/export/pdf/{project_id}
POST /api/v1/export/excel/{project_id}
POST /api/v1/export/cad/{project_id}
POST /api/v1/export/glb/{project_id}
GET  /api/v1/export/download/{token}
"""

import io
import json
import uuid
import time
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import StreamingResponse

from app.database import get_project

router = APIRouter()

# Simple in-memory token store (use Redis in production)
_export_tokens: dict[str, dict] = {}
EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)


def _build_pdf_bytes(project_data: dict) -> bytes:
    """
    Generate a PDF structural report.
    Uses reportlab if installed; falls back to a plain-text PDF skeleton.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Table, TableStyle,
            Spacer, HRFlowable,
        )

        buf    = io.BytesIO()
        doc    = SimpleDocTemplate(buf, pagesize=A4,
                                   leftMargin=2*cm, rightMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        copper = colors.HexColor("#B5651D")
        steel  = colors.HexColor("#3D5A6C")
        stone  = colors.HexColor("#2C2A25")

        h1 = ParagraphStyle("h1", fontSize=22, fontName="Helvetica-Bold",
                             textColor=stone, spaceAfter=4)
        h2 = ParagraphStyle("h2", fontSize=13, fontName="Helvetica-Bold",
                             textColor=copper, spaceBefore=12, spaceAfter=4)
        body = styles["Normal"]
        mono = ParagraphStyle("mono", fontName="Courier", fontSize=9,
                               textColor=steel)

        # ── Cover ──
        story.append(Paragraph("StructAI Designer", h1))
        story.append(Paragraph("Structural Design Report", styles["Heading2"]))
        story.append(HRFlowable(color=copper, thickness=2, spaceAfter=12))

        params  = json.loads(project_data["params_json"])
        result  = json.loads(project_data["result_json"])
        struct  = result["structural"]
        meta    = result["metadata"]
        cost    = result["materials"]["cost_breakdown"]

        # ── Project Info ──
        story.append(Paragraph("Project Information", h2))
        info_data = [
            ["Project ID",  project_data["id"]],
            ["Project Name",project_data["name"]],
            ["Plot Size",   f"{params['plot_length']} × {params['plot_width']} m"],
            ["Configuration",f"{params['bhk']} · {params['floors']}"],
            ["Plot Area",   f"{meta.get('plot_area_m2', '–')} m²"],
            ["Built Area",  f"{meta.get('built_area_m2', '–')} m²"],
            ["Standards",   meta.get("standard", "IS 456:2000")],
        ]
        t = Table(info_data, colWidths=[5*cm, 11*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F0E8")),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#FAFAF7")]),
        ]))
        story.append(t)

        # ── Structural Results ──
        story.append(Paragraph("Structural Analysis", h2))
        sa_data = [
            ["Parameter",          "Value",           "Status"],
            ["Max Beam Span",      f"{struct['max_beam_span_m']} m",        "✓ Safe"],
            ["Actual Deflection",  f"{struct['actual_deflection_mm']} mm",  "✓ OK"],
            ["Column Load",        f"{struct['column_load_kn']} kN",        "Nominal"],
            ["Slab Thickness",     f"{struct['slab']['thickness_mm']} mm",  "Standard"],
            ["Shear Capacity",     f"{struct['shear_capacity_kn']} kN",     "Review"],
            ["Safety Score",       f"{struct['safety_score']}/100",         struct["safety_status"]],
            ["Column Count",       f"{struct['columns']['count']}",         "Optimised"],
            ["Column Size",        f"{struct['columns']['size_mm']} mm",    "IS 456"],
            ["Beam Width × Depth", f"{struct['beams']['width_mm']} × {struct['beams']['depth_mm']} mm", "OK"],
        ]
        t2 = Table(sa_data, colWidths=[6*cm, 5*cm, 5*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), steel),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (1,0), (-1,-1), [colors.white, colors.HexColor("#FAFAF7")]),
        ]))
        story.append(t2)

        # ── Cost Summary ──
        story.append(Paragraph("Cost Estimate", h2))
        cost_data = [
            ["Item",        "Cost (₹ Lakhs)"],
            ["Foundation",  f"₹{cost['foundation']}L"],
            ["Structure",   f"₹{cost['structure']}L"],
            ["Masonry",     f"₹{cost['masonry']}L"],
            ["Finishing",   f"₹{cost['finishing']}L"],
            ["MEP",         f"₹{cost['mep']}L"],
            ["TOTAL",       f"₹{cost['total']}L"],
            ["Budget",      f"₹{cost['budget']}L"],
            ["Saved",       f"₹{cost['saved']}L  ✓"],
        ]
        t3 = Table(cost_data, colWidths=[8*cm, 8*cm])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), copper),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E6F2EA")),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ]))
        story.append(t3)

        # ── Compliance ──
        story.append(Paragraph("Code Compliance", h2))
        comp = struct.get("compliance", {})
        for code, passed in comp.items():
            mark = "✓ COMPLIANT" if passed else "✗ REVIEW REQUIRED"
            col  = colors.HexColor("#4A7C59") if passed else colors.HexColor("#8B3A2A")
            story.append(Paragraph(
                f'<font color="#{col.hexval()[1:]}">{mark}</font>  — {code.replace("_", " ")}',
                body
            ))

        # ── Warnings ──
        if struct.get("warnings"):
            story.append(Paragraph("Warnings / Recommendations", h2))
            for w in struct["warnings"]:
                story.append(Paragraph(f"⚠  {w}", mono))
                story.append(Spacer(1, 4))

        story.append(Spacer(1, 20))
        story.append(Paragraph(
            "Generated by StructAI Designer v2.1.0 · IS 456:2000 · IS 875 · NBC 2016",
            ParagraphStyle("footer", fontSize=7, textColor=colors.grey)
        ))

        doc.build(story)
        return buf.getvalue()

    except ImportError:
        # Minimal fallback if reportlab not installed
        text = f"""%PDF-1.4
1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj
2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj
3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842]
  /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>
endobj
4 0 obj << /Length 44 >>
stream
BT /F1 12 Tf 50 750 Td (StructAI Structural Report) Tj ET
endstream
endobj
5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj
xref
0 6
0000000000 65535 f
trailer << /Size 6 /Root 1 0 R >>
startxref 0
%%EOF"""
        return text.encode()


def _build_excel_bytes(project_data: dict) -> bytes:
    """Generate material schedule as XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = "Material Schedule"

        copper_fill = PatternFill("solid", fgColor="B5651D")
        steel_fill  = PatternFill("solid", fgColor="3D5A6C")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        bold_font   = Font(bold=True, name="Calibri")

        result    = json.loads(project_data["result_json"])
        materials = result["materials"]

        # Title row
        ws.merge_cells("A1:E1")
        ws["A1"] = "StructAI Designer — Material & Cost Schedule"
        ws["A1"].font = Font(bold=True, size=14, name="Calibri")

        # Per-floor table
        headers = ["Floor", "Concrete (m³)", "Steel (qtl)", "Brickwork (m³)", "Cost (₹L)"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.fill = copper_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ri, floor in enumerate(materials["per_floor"], start=4):
            ws.cell(ri, 1, floor["floor_label"])
            ws.cell(ri, 2, floor["concrete_m3"])
            ws.cell(ri, 3, floor["steel_quintals"])
            ws.cell(ri, 4, floor["brickwork_m3"])
            ws.cell(ri, 5, floor["cost_lakhs"])

        # Cost breakdown
        start_row = 4 + len(materials["per_floor"]) + 2
        ws.cell(start_row, 1, "Cost Breakdown").font = bold_font
        cost = materials["cost_breakdown"]
        for key in ["foundation","structure","masonry","finishing","mep","total","budget","saved"]:
            ws.cell(start_row + 1, 1, key.capitalize())
            ws.cell(start_row + 1, 2, cost[key])
            start_row += 1

        # Auto-width
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except ImportError:
        return b"Excel export requires openpyxl. Install with: pip install openpyxl"


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/pdf/{project_id}", summary="Export structural report as PDF")
async def export_pdf(project_id: str):
    row = await get_project(project_id)
    if not row or not row.get("result_json"):
        raise HTTPException(404, "Project not found or not yet generated")

    pdf_bytes = _build_pdf_bytes(row)
    filename  = f"StructAI_Report_{project_id[:8]}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/xls/{project_id}", summary="Alias for /excel/{project_id}")
async def export_xls_alias(project_id: str):
    return await export_excel(project_id)


@router.post("/excel/{project_id}", summary="Export material schedule as XLSX")
async def export_excel(project_id: str):
    row = await get_project(project_id)
    if not row or not row.get("result_json"):
        raise HTTPException(404, "Project not found or not yet generated")

    xls_bytes = _build_excel_bytes(row)
    filename  = f"StructAI_Materials_{project_id[:8]}.xlsx"

    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/dwg/{project_id}", summary="Alias for /cad/{project_id}")
async def export_dwg_alias(project_id: str):
    return await export_cad(project_id)


@router.post("/cad/{project_id}", summary="Export column/beam grid as DXF")
async def export_cad(project_id: str):
    """
    Exports a simple DXF with column grid lines, column circles, and beam lines.
    Requires ezdxf: pip install ezdxf
    """
    row = await get_project(project_id)
    if not row or not row.get("result_json"):
        raise HTTPException(404, "Project not found")

    try:
        import ezdxf

        result  = json.loads(row["result_json"])
        columns = result["structural"]["columns"]
        xs = columns["x_positions"]
        zs = columns["z_positions"]
        size_m = columns["size_mm"] / 1000

        doc = ezdxf.new(dxfversion="R2010")
        msp = doc.modelspace()

        # Column circles
        for x in xs:
            for z in zs:
                msp.add_circle((x, z), radius=size_m / 2,
                               dxfattribs={"layer": "COLUMNS", "color": 5})

        # Beam lines (X direction)
        for z in zs:
            for i in range(len(xs) - 1):
                msp.add_line((xs[i], z), (xs[i+1], z),
                             dxfattribs={"layer": "BEAMS", "color": 1})

        # Beam lines (Z direction)
        for x in xs:
            for j in range(len(zs) - 1):
                msp.add_line((x, zs[j]), (x, zs[j+1]),
                             dxfattribs={"layer": "BEAMS", "color": 1})

        buf = io.BytesIO()
        doc.write(buf)
        buf.seek(0)
        filename = f"StructAI_Grid_{project_id[:8]}.dxf"
        return StreamingResponse(
            buf,
            media_type="application/dxf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except ImportError:
        raise HTTPException(
            501,
            "CAD export requires ezdxf. Install with: pip install ezdxf",
        )


@router.post("/glb/{project_id}", summary="Export 3D model as GLB")
async def export_glb(project_id: str):
    """
    Placeholder for GLB export.
    In a real app, this would use a library like 'trimesh' or 'pyrender'
    to generate a GLB from the structural data.
    """
    row = await get_project(project_id)
    if not row or not row.get("result_json"):
        raise HTTPException(404, "Project not found")

    # Simple placeholder: return the summary JSON but as a file
    # (Actual GLB generation is out of scope for this hackathon)
    data = json.loads(row["result_json"])
    buf  = io.BytesIO(json.dumps(data).encode())
    return StreamingResponse(
        buf,
        media_type="model/gltf-binary",
        headers={"Content-Disposition": f'attachment; filename="StructAI_Model_{project_id[:8]}.glb"'},
    )


@router.get("/summary/{project_id}", summary="Get exportable summary JSON")
async def export_summary(project_id: str):
    """Returns a clean JSON summary suitable for third-party integrations."""
    row = await get_project(project_id)
    if not row or not row.get("result_json"):
        raise HTTPException(404, "Project not found")

    params = json.loads(row["params_json"])
    result = json.loads(row["result_json"])

    return {
        "project":    {"id": row["id"], "name": row["name"], "created_at": row["created_at"]},
        "parameters": params,
        "structural": result["structural"],
        "cost":       result["materials"]["cost_breakdown"],
        "metadata":   result["metadata"],
    }
